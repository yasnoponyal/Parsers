import json
import requests
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter



URL = 'https://dodopizza.ru/api/v5/menu?languageCode=ru-RU&countryCode=643&menuType=Delivery&locationId=000D3A38A8A3BCE911E9C7D3ED392E56'

import os
from dotenv import load_dotenv
load_dotenv()

COOKIES = {k: v for k, v in (pair.split("=", 1) for pair in os.getenv("DODO_COOKIES", "").split("; ") if "=" in pair)}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:148.0) Gecko/20100101 Firefox/148.0",
    "Accept": "*/*",
    "Accept-Language": "ru-RU,ru;q=0.9",
    "Referer": "https://dodopizza.ru/peterburg",
    "X-Requested-With": "XMLHttpRequest",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
}

FIELDNAMES   = ["Название", "Размер", "Тесто", "Цена", "Вес (г)", "Калории (на 100г)", "Общий каллораж", "Белки", "Жиры", "Углеводы", "Состав"]
SHEET_COLORS = {"Пиццы": "FFD700", "Другие блюда": "90EE90", "Напитки": "87CEEB", "Остальное": "DDA0DD"}

def total_cal(weight, calories):
    try:
        return round((float(weight) / 100) * float(calories), 1)
    except (TypeError, ValueError):
        return "?"



def load_data():
    try:
        with open("menu.json", encoding="utf-8") as f:
            print("Загружаю данные из menu.json...")
            return json.load(f)
    except FileNotFoundError:
        pass

    print("Файл menu.json не найден, пробую запрос к API...")
    try:
        r = requests.get(URL, headers=HEADERS, cookies=COOKIES, timeout=15)
        r.raise_for_status()
        with open("menu.json", "w", encoding="utf-8") as f:
            json.dump(r.json(), f, ensure_ascii=False, indent=2)
        print("Данные получены и сохранены в menu.json")
        return r.json()
    except Exception as e:
        print(f"Ошибка запроса: {e}")
        return None


def classify(item):
    variations = item.get('variations', [])
    traits = variations[0].get('product', {}).get('traits', {}) if variations else {}
    if traits.get('pizza'): return "Пиццы"
    if traits.get('drink'): return "Напитки"
    if traits.get('food'):  return "Другие блюда"
    return "Остальное"

def get_rows(item):
    name        = item.get('name', 'Без названия')
    description = (item.get('description') or '').replace('\n', ' ')
    variations  = item.get('variations', [])
    rows        = []

    is_pizza = variations and variations[0].get('product', {}).get('traits', {}).get('pizza', False)

    if is_pizza:
        size_groups = defaultdict(list)
        for var in variations:
            prod = var.get('product', {})
            size_groups[prod.get('size', '')].append(prod)

        DOUGH = ['Традиционное', 'Тонкое']
        for size, products in size_groups.items():
            for i, product in enumerate(products):
                food  = product.get('foodValue') or {}
                price = product.get('price')
                rows.append([
                    name, size, DOUGH[i] if i < 2 else f"Тесто {i+1}",
                    int(price) if isinstance(price, (int, float)) and price else "Уточняйте",
                    food.get('weight', '?'), food.get('calories', '?'),
                    total_cal(food.get('weight'), food.get('calories')),
                    food.get('proteins', '?'), food.get('fats', '?'),
                    food.get('carbohydrates', '?'), description,
                ])
    else:
        for var in variations:
            product = var.get('product', {})
            food    = product.get('foodValue') or {}
            price   = product.get('price')
            size    = product.get('sizeName') or product.get('size') or ''
            rows.append([
                name, size, '',
                int(price) if isinstance(price, (int, float)) and price else "Уточняйте",
                food.get('weight', '?'), food.get('calories', '?'),
                total_cal(food.get('weight'), food.get('calories')),
                food.get('proteins', '?'), food.get('fats', '?'),
                food.get('carbohydrates', '?'), description,
            ])
    return rows



def style_sheet(ws, header_color):
    col_widths  = [30, 12, 16, 8, 8, 12, 14, 8, 8, 10, 50]
    header_fill = PatternFill("solid", start_color=header_color)
    header_font = Font(name="Arial", bold=True, size=11)

    for col, (cell, width) in enumerate(zip(ws[1], col_widths), 1):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    row_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = row_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes        = "A2"
    ws.auto_filter.ref     = ws.dimensions
    ws.row_dimensions[1].height = 30


def main():
    data = load_data()
    if not data:
        return

    sheets_data = {"Пиццы": [], "Другие блюда": [], "Напитки": [], "Остальное": []}
    for item in data.get('items', []):
        sheets_data[classify(item)].extend(get_rows(item))

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(FIELDNAMES)
        for row in rows:
            ws.append(row)
        style_sheet(ws, SHEET_COLORS[sheet_name])
        print(f"  {sheet_name}: {len(rows)} строк")

    # from datetime import datetime
    filename = f"dodo_menu.xlsx"
    wb.save(filename)
    print(f"\nГотово! Файл {filename} сохранён.")

main()